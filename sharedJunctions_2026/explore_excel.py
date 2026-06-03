#!/usr/bin/env python3
import openpyxl
import pandas as pd

xlsx_path = '/gpfs0/tals/projects/Analysis/human_mouse_exons/ensembl115/sharedJunctions_2026/leafcutter_GSE115736_GSE116177/unique_sig_clusters_HN6.xlsx'

wb = openpyxl.load_workbook(xlsx_path)
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Dimensions: {ws.dimensions}')
    # Print first 5 rows
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        print(f'Row {i+1}: {row}')

# Focus on share_2_articles sheet
if 'share_2_articles' in wb.sheetnames:
    ws = wb['share_2_articles']
    print('\n\n=== DETAILED: share_2_articles ===')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    # Print header row
    header = [cell.value for cell in ws[1]]
    print('Headers:', header)
    # Print first 10 rows
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        print(f'Row {i+1}: {row}')
    
    # Show unique values in columns D-J for a few rows
    print('\nColumn D header:', ws.cell(1, 4).value)
    print('Column E header:', ws.cell(1, 5).value)
    print('Column F header:', ws.cell(1, 6).value)
    print('Column G header:', ws.cell(1, 7).value)
    print('Column H header:', ws.cell(1, 8).value)
    print('Column I header:', ws.cell(1, 9).value)
    print('Column J header:', ws.cell(1, 10).value)
    
    # Check unique values in col D-J
    for col_idx in range(4, 11):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        values = set()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx, values_only=True):
            if row[0] is not None:
                values.add(row[0])
        print(f'Col {col_letter} unique values: {sorted(values)[:20]}')
