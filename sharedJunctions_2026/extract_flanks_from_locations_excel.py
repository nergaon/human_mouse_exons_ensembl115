#!/usr/bin/env python3
"""Add left/right genomic flank sequences to an Excel file from chr:start:end locations.

Default behavior:
- Read genomic locations from column A of the first worksheet
- Fetch 11 bp immediately left of start plus the start base, and the end base plus 11 bp immediately right of end
- Write a new Excel file with two added columns

Coordinates are interpreted as 1-based inclusive.
Left flank: [start-11, start]
Right flank: [end, end+11]
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT_XLSX = Path(
    "/gpfs0/tals/projects/Analysis/human_mouse_exons/ensembl115/sharedJunctions_2026/"
    "leafcutter_GSE115736_GSE116177/clusters_sum_table_HN6.xlsx"
)
DEFAULT_FASTA = Path("/gpfs0/tals/projects/data/Genomes/hg38/grch38/GRCh38.primary_assembly.genome.fa")

LOCATION_RE = re.compile(r"^\s*([^:\s]+):(\d+):(\d+)\s*$")


@dataclass(frozen=True)
class FaiEntry:
    length: int
    offset: int
    line_bases: int
    line_width: int


class IndexedFasta:
    """Minimal random-access FASTA reader using a .fai index."""

    def __init__(self, fasta_path: Path):
        self.fasta_path = fasta_path
        self.fai_path = Path(str(fasta_path) + ".fai")
        if not fasta_path.exists():
            raise FileNotFoundError(f"FASTA not found: {fasta_path}")
        if not self.fai_path.exists():
            raise FileNotFoundError(f"FAI index not found: {self.fai_path}")
        self.index = self._load_fai(self.fai_path)
        self._fh = fasta_path.open("rb")

    @staticmethod
    def _load_fai(fai_path: Path) -> dict[str, FaiEntry]:
        out: dict[str, FaiEntry] = {}
        with fai_path.open("rt", encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                parts = line.rstrip("\n").split("\t")
                if len(parts) < 5:
                    continue
                name, length, offset, line_bases, line_width = parts[:5]
                out[name] = FaiEntry(
                    length=int(length),
                    offset=int(offset),
                    line_bases=int(line_bases),
                    line_width=int(line_width),
                )
        if not out:
            raise ValueError(f"No entries loaded from FAI: {fai_path}")
        return out

    def close(self) -> None:
        self._fh.close()

    def resolve_contig(self, contig: str) -> str | None:
        if contig in self.index:
            return contig

        alt = None
        if contig.startswith("chr"):
            alt = contig[3:]
        else:
            alt = f"chr{contig}"
        if alt in self.index:
            return alt

        mt_aliases = {
            "chrM": "MT",
            "MT": "chrM",
            "M": "MT",
            "chrMT": "MT",
        }
        if contig in mt_aliases and mt_aliases[contig] in self.index:
            return mt_aliases[contig]

        return None

    def fetch_1based(self, contig: str, start_1: int, end_1: int) -> str:
        """Fetch sequence on 1-based inclusive coordinates, clamped to contig bounds."""
        resolved = self.resolve_contig(contig)
        if resolved is None:
            return ""

        ent = self.index[resolved]
        if end_1 < 1 or start_1 > ent.length:
            return ""

        s = max(1, start_1)
        e = min(ent.length, end_1)
        if e < s:
            return ""

        start0 = s - 1
        remaining = e - s + 1
        pos0 = start0
        chunks: list[str] = []

        while remaining > 0:
            line_idx = pos0 // ent.line_bases
            in_line = pos0 % ent.line_bases
            take = min(remaining, ent.line_bases - in_line)
            byte_pos = ent.offset + (line_idx * ent.line_width) + in_line
            self._fh.seek(byte_pos)
            chunks.append(self._fh.read(take).decode("ascii"))
            pos0 += take
            remaining -= take

        return "".join(chunks).upper()


def parse_location(value: object) -> tuple[str, int, int] | None:
    if value is None:
        return None
    m = LOCATION_RE.match(str(value))
    if not m:
        return None
    contig = m.group(1)
    start = int(m.group(2))
    end = int(m.group(3))
    if start <= 0 or end <= 0:
        return None
    if end < start:
        start, end = end, start
    return contig, start, end


def default_output_path(input_xlsx: Path) -> Path:
    return input_xlsx.with_name(f"{input_xlsx.stem}_with_flanks.xlsx")


def add_flanks_to_excel(
    input_xlsx: Path,
    fasta_path: Path,
    output_xlsx: Path,
    flank_bp: int = 12,
    source_col: int = 1,
    sheet_name: str | None = None,
) -> None:
    wb = load_workbook(input_xlsx)
    ws = wb[sheet_name] if sheet_name else wb.active

    left_header = f"left_{flank_bp}bp_inclusive_seq"
    right_header = f"right_{flank_bp}bp_inclusive_seq"

    left_col = ws.max_column + 1
    right_col = ws.max_column + 2
    ws.cell(row=1, column=left_col, value=left_header)
    ws.cell(row=1, column=right_col, value=right_header)

    fasta = IndexedFasta(fasta_path)
    try:
        n_rows = 0
        n_ok = 0
        n_bad = 0
        for r in range(2, ws.max_row + 1):
            n_rows += 1
            loc = ws.cell(row=r, column=source_col).value
            parsed = parse_location(loc)
            if parsed is None:
                ws.cell(row=r, column=left_col, value="")
                ws.cell(row=r, column=right_col, value="")
                n_bad += 1
                continue

            contig, start, end = parsed
            left_seq = fasta.fetch_1based(contig, start - flank_bp, start)
            right_seq = fasta.fetch_1based(contig, end, end + flank_bp)
            ws.cell(row=r, column=left_col, value=left_seq)
            ws.cell(row=r, column=right_col, value=right_seq)
            n_ok += 1

    finally:
        fasta.close()

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    print(f"Input: {input_xlsx}")
    print(f"Genome FASTA: {fasta_path}")
    print(f"Sheet: {ws.title}")
    print(f"Rows processed: {n_rows}")
    print(f"Parsed locations: {n_ok}")
    print(f"Invalid/empty locations: {n_bad}")
    print(f"Output: {output_xlsx}")


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", type=Path, default=DEFAULT_INPUT_XLSX, help="Input .xlsx file")
    p.add_argument("--fasta", type=Path, default=DEFAULT_FASTA, help="Reference genome FASTA (.fai required)")
    p.add_argument("--output", type=Path, default=None, help="Output .xlsx path (default: *_with_flanks.xlsx)")
    p.add_argument("--sheet", type=str, default=None, help="Worksheet name (default: first sheet)")
    p.add_argument("--col", type=int, default=1, help="1-based source column index with chr:start:end")
    p.add_argument("--flank", type=int, default=12, help="Flank size in bp")
    p.add_argument(
        "--in-place",
        action="store_true",
        help="Overwrite input file instead of writing a new output file",
    )
    return p


def main() -> None:
    args = build_arg_parser().parse_args()
    input_xlsx = args.input
    output_xlsx = input_xlsx if args.in_place else (args.output if args.output else default_output_path(input_xlsx))

    if args.flank < 1:
        raise ValueError("--flank must be >= 1")
    if args.col < 1:
        raise ValueError("--col must be >= 1")

    add_flanks_to_excel(
        input_xlsx=input_xlsx,
        fasta_path=args.fasta,
        output_xlsx=output_xlsx,
        flank_bp=args.flank,
        source_col=args.col,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()
